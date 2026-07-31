# BrosLMV - Botones personalizados para CONTPAQi Comercial PRO
# Copyright (C) 2026 Cristofer Candelas Garcia
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

from __future__ import annotations

from typing import Any

from . import _bridge


class _Erp:
    """Espejo de ctx.erp (ErpContext de C#). Cada atributo es un método de XEngine que se
    relaya al addon en proceso (que tiene la conexión/engine vivos). Los nombres son los
    mismos que en C# (PascalCase): ctx.erp.GetProductStock(125, 0), ctx.erp.GetSalePrice(1)...
    El addon resuelve el método contra ErpContext (sin distinguir mayúsculas) y, si no es un
    wrapper tipado, cae a ctx.erp.Call(nombre, *args)."""

    def __getattr__(self, name: str):
        def _call(*args: Any) -> Any:
            return _bridge.call("erp", name, list(args))
        return _call


def _sql_lit(v: Any) -> str:
    """Convierte un valor Python a literal SQL seguro (escapa comillas)."""
    import datetime
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (datetime.date, datetime.datetime)):
        return "'" + v.isoformat(sep=" ") + "'"
    return "N'" + str(v).replace("'", "''") + "'"


# Caché por tabla de la columna identidad (PK autoincremental). Evita re-consultar.
_PK_CACHE: dict[str, str | None] = {}
# Caché del flag de solo-lectura del contexto (no cambia durante una ejecución).
_RO_CACHE: dict[str, bool] = {}


class _Record:
    """Registro 'active-record' genérico sobre una tabla. Patrón:

        doc = ctx.nuevo("docDocument")
        doc["ModuleID"] = 183
        doc.guardar()          # INSERT; devuelve el nuevo ID (también en doc.id)
        print(doc["DocumentID"])

    Reusa el relay SQL (conexión viva del addon). Genérico para cualquier tabla.
    """

    def __init__(self, ctx: "Context", tabla: str):
        self._ctx = ctx
        self._tabla = tabla
        self._campos: dict[str, Any] = {}
        self._originales: dict[str, Any] = {}
        self._pkcol: str | None = None
        self.id: Any = None

    def __setitem__(self, k: str, v: Any) -> None:
        self._campos[k] = v

    def __getitem__(self, k: str) -> Any:
        return self._campos.get(k)

    def set(self, **kwargs: Any) -> "_Record":
        self._campos.update(kwargs)
        return self

    def _guarda_lectura(self) -> None:
        if "ro" not in _RO_CACHE:
            try:
                _RO_CACHE["ro"] = bool(self._ctx.context().get("solo_lectura", False))
            except Exception:
                _RO_CACHE["ro"] = False
        if _RO_CACHE["ro"]:
            raise RuntimeError("Modo SOLO LECTURA activo: no se puede escribir.")

    def _pk(self) -> str | None:
        if self._tabla not in _PK_CACHE:
            try:
                _PK_CACHE[self._tabla] = self._ctx.scalar(
                    "SELECT name FROM sys.columns WHERE object_id=OBJECT_ID(%s) AND is_identity=1"
                    % _sql_lit(self._tabla))
            except Exception:
                _PK_CACHE[self._tabla] = None
        return _PK_CACHE[self._tabla]

    def guardar(self) -> Any:
        """INSERT de los campos seteados. Devuelve el nuevo ID (columna identidad)."""
        self._guarda_lectura()
        if not self._campos:
            raise ValueError("No hay campos que guardar.")
        cols = list(self._campos.keys())
        vals = ", ".join(_sql_lit(self._campos[c]) for c in cols)
        self._pkcol = self._pk()
        out = (" OUTPUT INSERTED.%s" % self._pkcol) if self._pkcol else ""
        sql = "INSERT INTO %s (%s)%s VALUES (%s)" % (self._tabla, ", ".join(cols), out, vals)
        if self._pkcol:
            self.id = self._ctx.scalar(sql)
            self._campos[self._pkcol] = self.id
        else:
            self._ctx.execute(sql)
            self.id = None
        self._originales = dict(self._campos)
        return self.id

    def actualizar(self) -> int:
        """UPDATE solo de los campos modificados (excepto la PK) por la PK ya conocida."""
        self._guarda_lectura()
        if not self._pkcol or self.id is None:
            raise RuntimeError("actualizar() requiere un registro ya guardado o cargado (con PK).")
        cambios = {c: v for c, v in self._campos.items()
                   if c != self._pkcol and (c not in self._originales or self._originales[c] != v)}
        if not cambios:
            return 0
        sets = ", ".join("%s=%s" % (c, _sql_lit(v)) for c, v in cambios.items())
        filas = self._ctx.execute("UPDATE %s SET %s WHERE %s=%s" %
                                  (self._tabla, sets, self._pkcol, _sql_lit(self.id)))
        self._originales.update(cambios)
        return filas

    def _cargar(self, pk_value: Any) -> "_Record":
        self._pkcol = self._pk()
        if not self._pkcol:
            raise RuntimeError("La tabla '%s' no tiene columna identidad." % self._tabla)
        rows = self._ctx.query(
            "SELECT TOP 1 * FROM %s WHERE %s=%s AND DeletedOn IS NULL"
            % (self._tabla, self._pkcol, _sql_lit(pk_value)))
        if not rows:
            raise RuntimeError("No se encontró registro en '%s' con %s=%s."
                               % (self._tabla, self._pkcol, pk_value))
        self._campos = dict(rows[0])
        self._originales = dict(rows[0])
        self.id = pk_value
        return self

    def eliminar(self) -> int:
        """Borrado LÓGICO (DeletedOn=GETDATE()) por la PK conocida."""
        self._guarda_lectura()
        if not self._pkcol or self.id is None:
            raise RuntimeError("eliminar() requiere un registro ya guardado (con PK).")
        return self._ctx.execute("UPDATE %s SET DeletedOn=GETDATE() WHERE %s=%s" %
                                 (self._tabla, self._pkcol, _sql_lit(self.id)))


class Context:
    @property
    def user_id(self) -> int:
        return int(_bridge.call("context").get("user_id", 0))

    @property
    def module_id(self) -> int:
        return int(_bridge.call("context").get("module_id", 0))

    @property
    def empresa(self) -> str:
        return str(_bridge.call("context").get("empresa", ""))

    @property
    def app_key(self) -> str:
        return str(_bridge.call("context").get("app_key", ""))

    def context(self) -> dict[str, Any]:
        return dict(_bridge.call("context"))

    @property
    def fila(self) -> dict[str, Any]:
        return dict(_bridge.call("context").get("fila_activa", {}))

    def get_selected_ids(self) -> list[int]:
        return [int(x) for x in (_bridge.call("get_selected_ids") or [])]

    def msg(self, text: str, title: str = "BrosLMV") -> None:
        _bridge.call("msg", text, title)

    def confirm(self, text: str, title: str = "Confirmar") -> bool:
        """Pregunta Sí/No al usuario. Bloquea hasta que responda.

        Ejemplo:
            if ctx.confirm("¿Timbrar el documento " + str(doc) + "?"):
                ctx.erp.Timbrar(doc)
        """
        return bool(_bridge.call("confirm", text, title))

    def form(self, spec: dict[str, Any]) -> dict[str, Any]:
        """Muestra un formulario WinForms renderizado por el addon.

        Ejemplo:
            r = ctx.form({
                "title": "Datos",
                "fields": [
                    {"name": "proveedor", "label": "Proveedor", "type": "text", "required": True},
                    {"name": "cantidad", "label": "Cantidad", "type": "decimal", "default": 1},
                ],
            })
            if r["submitted"]:
                print(r["values"])

        Grid editable opcional (partidas/renglones), con "grid": {...} en el spec:
            r = ctx.form({
                "title": "Partidas",
                "grid": {
                    "columns": [
                        {"name": "producto", "caption": "Producto", "type": "text"},
                        {"name": "cantidad", "caption": "Cantidad", "type": "decimal", "editable": True},
                    ],
                    "rows": [{"producto": "Tornillo", "cantidad": 10}],
                    "allow_add": True,
                    "allow_delete": True,
                },
            })
            if r["submitted"]:
                for fila in r["grid_rows"]:
                    print(fila["producto"], fila["cantidad"])
        """
        return dict(_bridge.call("form", spec or {}))

    def show_html(self, html: str, title: str = "BrosLMV", width: int = 800,
                   height: int = 600, modal: bool = True) -> None:
        """Muestra una ventana con HTML/CSS/JS renderizado por WebView2 (Edge/Chromium),
        embebida en el addon dentro del proceso de Comercial.

        Ejemplo:
            ctx.show_html("<h1>Hola</h1><p>Reporte generado.</p>", title="Reporte", width=900, height=650)
        """
        _bridge.call("show_html", html or "", title, width, height, modal)

    def show_html_formulario(self, html: str, title: str = "BrosLMV", width: int = 900,
                              height: int = 700, timeout_ms: int = 600000) -> dict[str, Any]:
        """Como show_html, pero BLOQUEA el script hasta que la pagina envie datos de vuelta
        o el usuario cierre la ventana. Para formularios HTML reales (crear/guardar algo),
        a diferencia de show_html que es de una sola via (solo para mostrar).

        En el HTML, el boton de guardar/enviar debe llamar:
            window.chrome.webview.postMessage(JSON.stringify({campo1: valor1, ...}))

        Regresa ese mismo diccionario, con "submitted" agregado (True si se envio algo,
        False si se cerro la ventana sin enviar nada).

        IMPORTANTE: si el formulario puede tardar mas de 2 minutos en llenarse (lo normal
        para un humano), agrega "# timeout: 1800" (segundos) en las primeras lineas del
        script -- sin eso, el timeout por default del script completo (no de esta ventana)
        puede tronar antes de que el usuario termine de llenar el formulario.

        Ejemplo:
            r = ctx.show_html_formulario("<input id='n'><button onclick=\"" +
                "window.chrome.webview.postMessage(JSON.stringify({nombre: n.value}))\">" +
                "Guardar</button>")
            if r["submitted"]:
                ctx.msg("Recibido: " + r["nombre"])
        """
        return dict(_bridge.call("show_html_formulario", html or "", title, width, height, timeout_ms))

    def dashboard(self, title: str, data: list[dict[str, Any]],
                   columns: list[dict[str, str]] | None = None,
                   width: int = 1000, height: int = 700, modal: bool = True) -> None:
        """Muestra un dashboard HTML completo (tabla ordenable, buscador, paginación y
        exportar a Excel) a partir de una lista de dict -- sin escribir HTML/CSS/JS.

        Usa la plantilla y las librerías compartidas del runtime (C:\\BrosLMV\\lib\\dashboard\\,
        las trae el instalador) -- NO copia nada a la carpeta del script, así que funciona
        igual en cualquier terminal donde esté instalado BrosLMV, sin distribuir nada por
        script ni por cliente. Ver MANUAL.md / DASHBOARDS_HTML.md.

        columns: opcional, [{"key": "Proveedor", "label": "Proveedor"}, ...]. Si se omite,
        las columnas se infieren de las llaves del primer registro de `data`.

        Los datos siempre viajan comprimidos (gzip+base64) -- por eso nunca choca con el
        límite real de ~2MB de WebView2 (NavigateToString), sin importar cuántas filas
        traiga el reporte.

        Ejemplo:
            filas = ctx.query("SELECT Proveedor, Total FROM ...")
            ctx.dashboard("Compras por proveedor", filas)
        """
        import base64
        import datetime
        import gzip
        import json
        import os

        plantilla_path = os.path.join(r"C:\BrosLMV\lib\dashboard", "dashboard_base.html")
        try:
            with open(plantilla_path, "r", encoding="utf-8") as f:
                html = f.read()
        except FileNotFoundError:
            raise RuntimeError(
                "No se encontró " + plantilla_path + " -- reinstala/actualiza BrosLMV "
                "(la trae el instalador, no viaja con el script)."
            )

        crudo = json.dumps(data or [], default=str).encode("utf-8")
        datos_b64 = base64.b64encode(gzip.compress(crudo, compresslevel=9)).decode("ascii")

        html = html.replace("{{TITLE_JSON}}", json.dumps(title or "BrosLMV"))
        html = html.replace("{{TITLE}}", title or "BrosLMV")
        html = html.replace("{{FECHA}}", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        html = html.replace("{{COLUMNS_JSON}}", json.dumps(columns or []))
        html = html.replace("{{DATA_B64_JSON}}", json.dumps(datos_b64))

        self.show_html(html, title=title, width=width, height=height, modal=modal)

    def read_excel(self, path: str, sheet: str | None = None) -> list[dict[str, Any]]:
        """Lee un archivo .xlsx y regresa una lista de dict (una por fila, claves = encabezados
        de la primera fila). No requiere Excel instalado (usa openpyxl, no automatiza Excel.exe).

        Ejemplo:
            filas = ctx.read_excel(r"C:\\Cotizaciones\\productos.xlsx")
            for fila in filas:
                print(fila["ProductID"], fila["Cantidad"])
        """
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb.active
        filas_iter = ws.iter_rows(values_only=True)
        try:
            encabezados = [str(h) if h is not None else "" for h in next(filas_iter)]
        except StopIteration:
            return []

        resultado: list[dict[str, Any]] = []
        for fila in filas_iter:
            if all(v is None for v in fila):
                continue
            resultado.append({encabezados[i]: fila[i] for i in range(min(len(encabezados), len(fila)))})
        return resultado

    def write_excel(self, rows: list[dict[str, Any]], path: str, sheet_name: str = "Hoja1") -> None:
        """Escribe una lista de dict a un archivo .xlsx (encabezados = claves del primer dict).
        No requiere Excel instalado.

        Ejemplo:
            ctx.write_excel([{"Producto": "Tornillo", "Cantidad": 10}], r"C:\\Reportes\\salida.xlsx")
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if rows:
            encabezados = list(rows[0].keys())
            ws.append(encabezados)
            for fila in rows:
                ws.append([fila.get(h) for h in encabezados])

        wb.save(path)

    def select_file(self, title: str = "Seleccionar archivo", filter: str = "Todos los archivos|*.*",
                     save: bool = False, initial_dir: str = "") -> str:
        """Abre un diálogo nativo de Windows para elegir un archivo (o dónde guardarlo, si
        save=True). Regresa la ruta elegida, o "" si el usuario canceló.

        Ejemplo:
            ruta = ctx.select_file("Elegir Excel de productos", "Excel|*.xlsx;*.xls")
            if ruta:
                filas = ctx.read_excel(ruta)
        """
        return str(_bridge.call("select_file", title, filter, save, initial_dir) or "")

    def select_folder(self, title: str = "Seleccionar carpeta", initial_dir: str = "") -> str:
        """Abre un diálogo nativo de Windows para elegir una carpeta. Regresa la ruta
        elegida, o "" si el usuario canceló."""
        return str(_bridge.call("select_folder", title, initial_dir) or "")

    def log(self, text: str, level: str = "INFO") -> None:
        # log(text) o log(level, text) segun cuantos argumentos lleguen al host.
        if level == "INFO":
            _bridge.call("log", text)
        else:
            _bridge.call("log", level, text)

    def progress(self, text: str = "", percent: int = 0) -> None:
        _bridge.call("progress", text, int(percent))

    @property
    def execution_id(self) -> str:
        return str(_bridge.call("context").get("execution_id", ""))

    def query(self, sql: str, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        return _bridge.call("query", sql, params or {})

    def scalar(self, sql: str, params: dict[str, Any] | None = None) -> Any:
        return _bridge.call("scalar", sql, params or {})

    def execute(self, sql: str, params: dict[str, Any] | None = None) -> int:
        return int(_bridge.call("execute", sql, params or {}))

    @property
    def erp(self) -> _Erp:
        return _Erp()

    def nuevo(self, tabla: str) -> _Record:
        """Crea un registro 'active-record' para INSERT en `tabla`:

            it = ctx.nuevo("docDocumentItem")
            it["DocumentID"] = doc_id
            it["ProductID"]  = 1
            it.guardar()
        """
        return _Record(self, tabla)


    def registro(self, tabla: str, pk: Any) -> _Record:
        """Carga un registro existente por su PK (identidad):

            doc = ctx.registro("docDocument", 11556)
            doc["Comments"] = "Modificado"
            doc.actualizar()   # solo envía los campos que cambiaron
        """
        return _Record(self, tabla)._cargar(pk)


ctx = Context()
